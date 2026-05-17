import random

from aiogram import Router, types, F, Bot
from aiogram.filters import Command
from aiogram.fsm.state import State, StatesGroup
from services.db import schedule
from services.db.user_group import check_user_group
from services import get_weeks
from aiogram.fsm.context import FSMContext
from random import choice as r_choice
import json, io
from pathlib import Path
from aiogram.exceptions import TelegramBadRequest
from utils import keyboards

router = Router()


class ScheduleState(StatesGroup):
    group_getting = State()

mrkp = None

codes = {}
try:
    config_path = Path(__file__).parent.parent.parent.parent / "config" / "example-time.json"
    with open(config_path, encoding="utf-8") as f:
        data_ = json.load(f)
            #print(data)
        for c in data_["Times"]:
                codes[c["Code"]] = (
                    c["TimeFrom"][-8:-3],
                    c["TimeTo"][-8:-3],
                )
except FileNotFoundError:
    print("example-time.json not found :(")

@router.message(Command("schedule"))
async def cmd_get_schedule(message: types.Message, bot: Bot):
    await _get_schedule_logic(message, message.from_user.id, bot)


@router.callback_query(F.data == "watch_schedule_btn")
async def btn_watch_schedule(callback: types.CallbackQuery, bot: Bot):
    await _get_schedule_logic(callback.message, callback.from_user.id, bot)
    await callback.answer()


async def _safe_edit_text(msg: types.Message, text: str, **kwargs):
    if msg.text == text:
        return msg
    try:
        return await msg.edit_text(text, **kwargs)
    except TelegramBadRequest as e:
        if "message is not modified" in str(e):
            return msg
        raise


async def _get_schedule_logic(message: types.Message, user_id: int, bot: Bot, weektype:int = 1, sent_message = None):
    if not sent_message:
        sent_message = await message.answer(
            r_choice([
                "Загружаю расписание. Ожидайте...",
                "Расписание отправляется...",
                "Получаем расписание...",
                "Секунду... Расписание обрабатывается",
                "Ваше расписание почти готово..."
        ])
        
    )
        
        s_data = True
    
    else:
        s_data = False

    res = await check_user_group(user_id)
    if res is None or "group_name" not in res.keys() or len(res["group_name"]) < 4:
        await _safe_edit_text(
            sent_message,
            text="Сначала вы должны зарегистрировать свою группу.\nИспользуйте команду /group"
        )
        return
    
    try:
        group_name = res.get("group_name")
        # response = schedule.Schedule(group_name=group_name).run_()
        response1, response2, response3 = await schedule.Schedule(group_name=group_name).get_schedule_async(prev_next=True)
        #print(63)
    except Exception as e:
        print("Schedule error:", e)
        response1, response2, response3 = None, None, None

    # print(80)
    # print(response)
    #print(codes)
    st = []
    stWeek = []
    
    if not response2:
        await _safe_edit_text(sent_message, "Пока что пусто")
        return
    
    #print([response1, response2, response3])
    for response in [response1, response2, response3]:
        string = """"""
        week_name, days_data = response
        if not days_data:
            await _safe_edit_text(sent_message, "Пока что пусто")
            return
        string += "("+week_name+")"
        stWeek.append(week_name)
        for day, contents in days_data.items():
            if len(contents) == 0:
                pass
            else:
                string += f"""\n\n━━━━━━━━━━━━━━━━━━\n📅 <b>{day}</b>"""
            for lesson in contents:
                time_code = int(lesson["time_code"])
                # print(100)
                time_range = codes.get(time_code, ("", ""))
                # print(time_range)
                string += (
                    f"\n\n<b>{lesson['time']}</b>"
                    f" {time_range[0]} - {time_range[1]}\n"
                )
                string += f"{lesson['subject']} ({lesson['room']})"
        st.append(string)

    if not st:
        await _safe_edit_text(sent_message, "Пока что пусто")
        return
    if message.message_id != sent_message.message_id:
        try:
            await bot.unpin_all_chat_messages(message.chat.id)
        except TelegramBadRequest:
            pass

        try:
            if s_data:
                await sent_message.pin()
        except TelegramBadRequest:
            pass
        
    else:
        pass
    
    # mapping_prev = {
    #     0:3,
    #     1:0,
    #     2:1,
    #     3:2
    # }
    
    # mapping_next = {
    #     0:1,
    #     1:2,
    #     2:3,
    #     3:0
    # }

    final_text = f"<b>Вот твое расписание</b>\n{st[weektype]}"
    mappingPR = {
        0: None,
        1: stWeek[0],
        2: stWeek[1]
    }
    
    mappingNX = {
        0: stWeek[1],
        1: stWeek[2],
        2: None
    }
    pr =  mappingPR.get(weektype)
    nx = mappingNX.get(weektype)
    
    # print(pr, nx)

    if weektype == 1:
        prev_cb = "swipe_week:0"
        next_cb = "swipe_week:2"
    elif weektype == 0:
        prev_cb = None
        next_cb = "swipe_week:1"
    else:
        prev_cb = "swipe_week:1"
        next_cb = None

    mrkp = keyboards.swipe_schedule_kb(pr, nx, prev_data=prev_cb or "prev_week", next_data=next_cb or "next_week")
    #print(stWeek)
    await _safe_edit_text(sent_message, final_text, parse_mode="HTML", reply_markup=mrkp)

@router.callback_query(F.data.startswith("swipe_week:"))
async def swipe_schedule(callback: types.CallbackQuery, bot: Bot):
    try:
        weektype = int(callback.data.split(":", 1)[1])
    except (ValueError, IndexError):
        await callback.answer()
        return
    await _get_schedule_logic(callback.message, callback.from_user.id, bot, weektype=weektype, sent_message=callback.message)
    M = {
        "0": "Предыдущая неделя",
        "1": "Текущая неделя",
        "2": "Следующая неделя"
    }
    await callback.answer(M.get(callback.data[-1]))

@router.callback_query(F.data == "prev_week")
async def go_prev_sc(callback: types.CallbackQuery, bot: Bot):
    await _get_schedule_logic(callback.message, callback.from_user.id, bot, weektype=0, sent_message=callback.message)

@router.callback_query(F.data == "next_week")
async def go_prev_sc(callback: types.CallbackQuery, bot: Bot):
    await _get_schedule_logic(callback.message, callback.from_user.id, bot, weektype=2, sent_message=callback.message)
    
@router.callback_query(F.data == "download_schedule")
async def give_info_days(callback: types.CallbackQuery, state: FSMContext):
    if state:
        data = await state.get_data()
        selected_days = data.get("selected_days", [])
            
    else:
        selected_days = None
    mk = keyboards.dwn_days_kb(get_weeks.get_range(), selected_days=selected_days)  
    ALLSTR = """"""
    for week in get_weeks.get_range():
        ALLSTR += "\n"
        for day in week:
            if week[-1] == day:
                ALLSTR += day
            else:
                ALLSTR += day + ", "
                
    group = await check_user_group(callback.from_user.id)
    text = f"""<b>Скачать расписание для группы {group.get("group_name") if group else ""}</b>\n\nВыберите дни (<i>ПН - СБ + ВСЕ сразу</i>): {ALLSTR}"""
    
    await _safe_edit_text(callback.message, text=text, parse_mode="HTML", reply_markup=mk)
   
@router.callback_query(F.data == "download_new_sc")
async def rem(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    data:list = data.get('selected_days') if data else "Ничего не выбрано"
    # print(data)
    if data != "Ничего не выбрано" and len(data) > 0:
        data = ', '.join(data)
        await callback.answer()
        group = await check_user_group(callback.from_user.id)
        text = f"""<b>Скачать расписание для группы {group.get("group_name") if group else ""}</b>\n\n{data}"""
        await _safe_edit_text(callback.message, text=text, parse_mode="HTML", reply_markup=keyboards.go_or_back_kb())
    
    elif data == "Ничего не выбрано":
        await callback.answer("Выберите дни для скачивания!")
    
    else:
        await callback.answer("Ошибка. Выберите дни для скачивания!")

@router.callback_query(F.data == "cancel_add")
async def cancel_add(callback: types.CallbackQuery, state: FSMContext, bot: Bot):
    if state:
        await state.clear()
    await _get_schedule_logic(callback.message, callback.from_user.id, bot=bot, sent_message=callback.message)
    await callback.answer("Отменяем скачивание...")
    
@router.callback_query(F.data.startswith("add_"))
async def handle_day_selection(callback: types.CallbackQuery, state: FSMContext):
    clicked_day = callback.data.replace("add_", "")
    await callback.answer()
    data = await state.get_data()
    selected_days = data.get("selected_days", [])
    
    if clicked_day in selected_days:
        selected_days.remove(clicked_day)
    else:
        selected_days.append(clicked_day)
    
    await state.update_data(selected_days=selected_days)

    new_kb = keyboards.dwn_days_kb(get_weeks.get_range(), selected_days=selected_days)
    
    await callback.message.edit_reply_markup(reply_markup=new_kb)

    
    
@router.callback_query(F.data.startswith("all_"))
async def handle_all_week_selection(callback: types.CallbackQuery, state: FSMContext):
    week_index = int(callback.data.split("_")[1])

    data = await state.get_data()
    selected_days = data.get("selected_days", [])
    
    target_week = get_weeks.get_range()[week_index]
    

    all_present = all(str(day) in selected_days for day in target_week)
    
    if all_present:
        for day in target_week:
            if str(day) in selected_days:
                selected_days.remove(str(day))
    else:
        for day in target_week:
            day_str = str(day)
            if day_str not in selected_days:
                selected_days.append(day_str)

    await state.update_data(selected_days=selected_days)
    new_kb = keyboards.dwn_days_kb(get_weeks.get_range(), selected_days=selected_days)
    await callback.message.edit_reply_markup(reply_markup=new_kb)
    await callback.answer()
    
@router.callback_query(F.data.startswith("schedule_"))
async def dwn(callback: types.CallbackQuery, state: FSMContext, bot: Bot):
    user_id = callback.from_user.id
    res = await check_user_group(user_id)
    done = False

    
    if res is None:
        await _safe_edit_text(
            callback.message,
            text="Сначала вы должны зарегистрировать свою группу.\nИспользуйте команду /group"
        )
        return
    
    try:
        group_name = res.get("group_name")
        # response = schedule.Schedule(group_name=group_name).run_()
        response1, response2, response3 = await schedule.Schedule(group_name=group_name).get_schedule_async(prev_next=True)
        #print(63)
    except Exception as e:
        print("Schedule error:", e)
        response1, response2, response3 = None, None, None
        
    all_dates = get_weeks.get_range()
    all_dates_clear = []
    # for week in all_dates:
    #     for day in week:
    #         all_dates_clear.append(day)
    st_HTML = []
    st_TXT = []
    resp_idx = 0  
    
    data = await state.get_data()
    req_days = data["selected_days"] if data else []
    
    await _get_schedule_logic(callback.message, callback.from_user.id, bot=bot, sent_message=callback.message)
        
    st_cfg = {} #html
    st_cfg_txt = {} #txt
    
    
    for idx, response in enumerate([response1, response2, response3]):
        
        current_week_by_dates = all_dates[idx]
        week_name, days_data = response

        st_cfg[week_name] = {}
        st_cfg_txt[week_name] = {}
        
        
        for day_idx, (dayN, dayC) in enumerate(days_data.items()):
            stringHTML = """"""
            stringTXT = """"""
            # if dayN == "Понедельник":
            #     stringHTML += "<div class='week'>"
            stringTXT += f"\t📆 {dayN} {current_week_by_dates[day_idx] if  day_idx < len(current_week_by_dates) else ""}\n"
            stringHTML += "<div class='day'>"
            stringHTML += f"<br><h2>{dayN}{"<br>"+ current_week_by_dates[day_idx] if  day_idx < len(current_week_by_dates) else ""}</h2>"
            
            if not dayC:
                stringHTML += "<p>Занятий нет</p>"
                stringTXT += "\nЗанятий нет\n"
            else:
                for lesson in dayC:
                    time_code = int(lesson["time_code"])
                    time_range = codes.get(time_code, ("", ""))

                    stringHTML += (
                        f"<h4>{lesson['time']}</h4>"
                        f"<h5>{time_range[0]} - {time_range[1]}</h5>"
                        f"<h6>{lesson['teacher']}</h6>"
                        f"<p>{lesson['subject']} <span>({lesson['room']})</span></p>"
                    )
                    
                    stringTXT += (
                        f"\n{lesson['time']}\n"
                        f"  {time_range[0]} - {time_range[1]}\n"
                        f"  {lesson['teacher']}\n"
                        f"  {lesson['subject']} ({lesson['room']})\n"
                    )
            stringHTML += "</div>"
            stringTXT += "\n"
            
            if day_idx < len(current_week_by_dates):
                date_key = current_week_by_dates[day_idx]
                st_cfg[week_name][date_key] = stringHTML
                st_cfg_txt[week_name][date_key] = stringTXT

        full_week_html = "<div class='week'>" + "".join(st_cfg[week_name].values()) + "</div>"
        full_week_txt = "\n" + "\n".join(st_cfg_txt[week_name].values())
        st_HTML.append(full_week_html)
        st_TXT.append(full_week_txt)

    need_data = """"""
    for week, data in st_cfg.items():
        need_data += "<div class='week'>"
        for key_day, day_data in data.items():
            for d in req_days:
                if d == key_day:
                    need_data += day_data
        need_data += "</div>"

    
    if callback.data.endswith("html"):
        await _safe_edit_text(callback.message, "<b>Финальный шаг!</b>\n\nВыберите темы:", parse_mode="HTML", reply_markup=keyboards.styles_choise_kb())
        
        
        
    if callback.data.endswith("txt"):
        path = io.BytesIO()
        path.write(full_week_txt.encode("utf-8"))
        path.seek(0)
        await callback.answer("Файл готов. Отправка...")
        await callback.message.answer_document(
            document=types.BufferedInputFile(
                path.getvalue(),
                filename=f"schedule_{group_name}.txt"
            ),
            caption=f"Расписание для {group_name}"
        )
        await callback.answer("Скачано!")
        done = True
    
    if callback.data.endswith("exel"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"

        week_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        week_font = Font(color="FFFFFF", bold=True, size=14)

        day_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        day_font = Font(color="1D8348", bold=True, size=12)
        
        lesson_font = Font(size=11)
        time_font = Font(bold=True, color="34495E")
        
        thin_border = Border(
            left=Side(style='thin', color="BDC3C7"),
            right=Side(style='thin', color="BDC3C7"),
            top=Side(style='thin', color="BDC3C7"),
            bottom=Side(style='thin', color="BDC3C7")
        )
        
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15

        current_row = 1


        for idx, response in enumerate([response1, response2, response3]):
            if not response: continue
            
            week_name, days_data = response
            current_week_dates = all_dates[idx]


            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1, value=week_name)
            cell.fill = week_fill
            cell.font = week_font
            cell.alignment = alignment_center
            current_row += 1

            for day_idx, (day_name, lessons) in enumerate(days_data.items()):
                date_str = current_week_dates[day_idx] if day_idx < len(current_week_dates) else ""
                

                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                day_cell = ws.cell(row=current_row, column=1, value=f"{day_name} ({date_str})")
                day_cell.fill = day_fill
                day_cell.font = day_font
                day_cell.alignment = alignment_left
                current_row += 1

                if not lessons:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
                    cell = ws.cell(row=current_row, column=1, value="Пар нет, отдыхаем!")
                    cell.alignment = alignment_center
                    cell.font = Font(italic=True, color="7F8C8D")
                    current_row += 1
                else:

                    headers = ["Время", "Предмет", "Преподаватель", "Ауд."]
                    for col_num, header in enumerate(headers, 1):
                        c = ws.cell(row=current_row, column=col_num, value=header)
                        c.font = Font(bold=True)
                        c.border = thin_border
                    current_row += 1

                    for lesson in lessons:
                        time_code = int(lesson["time_code"])
                        time_range = codes.get(time_code, ("??:??", "??:??"))
                        
                        data_row = [
                            f"{time_range[0]} - {time_range[1]}",
                            lesson['subject'],
                            lesson['teacher'],
                            lesson['room']
                        ]
                        
                        for col_num, value in enumerate(data_row, 1):
                            c = ws.cell(row=current_row, column=col_num, value=value)
                            c.border = thin_border
                            c.alignment = alignment_left
                            if col_num == 1: c.font = time_font
                        
                        current_row += 1
                
                current_row += 1

        path = io.BytesIO()
        wb.save(path)
        path.seek(0)

        await callback.answer("Excel файл готов!")
        await callback.message.answer_document(
            document=types.BufferedInputFile(
                path.getvalue(),
                filename=f"schedule_{group_name}.xlsx"
            ),
            caption=f"Расписание для {group_name}"
        )  
        await callback.answer("Скачано!")
        done = True
        
    if callback.data.endswith("img"):
        from PIL import Image, ImageDraw, ImageFont
        import textwrap
        import os
        await callback.answer("Рисую расписание...")

        (color1, color2, color3) = random.choice([
                ("#27AE60", "#1D8348", "#D5F5E3"),
                ("#279EAE", "#1D5283", "#D5E2F5"),
                ("#8827AE", "#5A1D83", "#EFD5F5"),
                ("#E2DF33", "#FAAE09", "#F5F5D5")
            ])

        try:
            font_path = os.path.join(os.path.dirname(__file__), "Roboto-Regular.ttf")
            font_title = ImageFont.truetype(font_path, 28)
            font_week = ImageFont.truetype(font_path, 22)
            font_day = ImageFont.truetype(font_path, 20)
            font_text = ImageFont.truetype(font_path, 18)
            font_small = ImageFont.truetype(font_path, 14)
        except OSError:
            font_title = font_week = font_day = font_text = font_small = ImageFont.load_default()

        responses = [response1, response2, response3]
        
        media_list = []
        
        for idx, response in enumerate(responses):
            if not response: continue
                
            week_name, days_data = response
            current_week_dates = all_dates[idx] if idx < len(all_dates) else []


            valid_days_to_draw = []
            for day_idx, (day_name, lessons) in enumerate(days_data.items()):
                date_str = current_week_dates[day_idx] if day_idx < len(current_week_dates) else ""
                if req_days and date_str not in req_days:
                    continue
                valid_days_to_draw.append((day_name, date_str, lessons))


            if not valid_days_to_draw:
                continue


            width = 1200 

            temp_img = Image.new('RGB', (width, 5000), color='#f4f7f2')
            draw = ImageDraw.Draw(temp_img)
            
            y = 20
            draw.text((20, y), f"Расписание: {group_name}", fill="#2d3436", font=font_title)
            y += 60

            draw.rounded_rectangle([20, y, width-20, y+45], radius=8, fill=color1)
            draw.text((40, y+10), str(week_name), fill="white", font=font_week)
            y += 70

            col_x = [20, 610]
            col_w = 570
            col_y = [y, y]
            
            for d_idx, (day_name, date_str, lessons) in enumerate(valid_days_to_draw):
                c_idx = d_idx % 2
                c_x = col_x[c_idx]

                draw.rounded_rectangle([c_x, col_y[c_idx], c_x + col_w, col_y[c_idx]+35], radius=6, fill=color3)
                draw.text((c_x + 20, col_y[c_idx] + 5), f"{day_name} ({date_str})", fill=color2, font=font_day)
                col_y[c_idx] += 45

                if not lessons:
                    draw.text((c_x + 20, col_y[c_idx]), "Пар нет", fill="#7F8C8D", font=font_text)
                    col_y[c_idx] += 40
                else:
                    for lesson in lessons:
                        t_code = int(lesson.get("time_code", 0))
                        t_range = codes.get(t_code, ("??:??", "??:??"))
                        time_str = f"· {t_range[0]} - {t_range[1]}"
                        
                        draw.text((c_x + 20, col_y[c_idx]), time_str, fill="#34495E", font=font_text)
                        col_y[c_idx] += 25
                        
                        subj = f"{lesson.get('subject', '---')} ({lesson.get('room', '-')})"
                        wrapped_subject = textwrap.wrap(subj, width=40) 
                        for line in wrapped_subject:
                            draw.text((c_x + 40, col_y[c_idx]), line, fill="#2d3436", font=font_text)
                            col_y[c_idx] += 25
                        
                        teacher = f"{lesson.get('teacher', '---')}"
                        draw.text((c_x + 40, col_y[c_idx]), teacher, fill="#7F8C8D", font=font_small)
                        col_y[c_idx] += 25
                        
                        draw.line([c_x + 20, col_y[c_idx], c_x + col_w - 20, col_y[c_idx]], fill="#BDC3C7", width=1)
                        col_y[c_idx] += 15
                
                col_y[c_idx] += 20 


            final_y = max(col_y[0], col_y[1]) + 20 
            week_img = temp_img.crop((0, 0, width, max(final_y, 300)))

            path = io.BytesIO()
            week_img.save(path, format='PNG')
            path.seek(0)
            file = types.BufferedInputFile(path.getvalue(), filename=f"week_{idx+1}.png")
            media_list.append(types.InputMediaPhoto(
                media=file,
                caption=f"📅 Расписание\nГруппа: <b>{group_name}</b>" if len(media_list) == 0 else "", 
                parse_mode="HTML"
            ))
            # await callback.message.answer_photo(
            #     photo=types.BufferedInputFile(path.getvalue(), filename=f"week_{idx+1}.png"),
            #     caption=f"📅 {week_name}\nГруппа: <b>{group_name}</b>",
            #     parse_mode="HTML"
            # )
        if media_list:
            await callback.message.answer_media_group(media=media_list)
            await callback.answer("Расписание готово!")
        else:
            await callback.answer("Нет данных для вывода", show_alert=True)

        await callback.answer("Готово!")
        done = True
    
    if done:
        if state:
            await state.clear()
        
@router.callback_query(F.data.startswith("html_"))
async def send_schedule(callback: types.CallbackQuery, state: FSMContext, bot: Bot):
    user_id = callback.from_user.id
    res = await check_user_group(user_id)
    done = False
    await _get_schedule_logic(callback.message, callback.from_user.id, bot=bot, sent_message=callback.message)
    all_dates = get_weeks.get_range()
    all_dates_clear = []
    
    data = await state.get_data()
    req_days = data["selected_days"] if data else []
    
    try:
        group_name = res.get("group_name")
        # response = schedule.Schedule(group_name=group_name).run_()
        response1, response2, response3 = await schedule.Schedule(group_name=group_name).get_schedule_async(prev_next=True)
        #print(63)
    except Exception as e:
        print("Schedule error:", e)
        response1, response2, response3 = None, None, None
    
    theme = callback.data.split("_")[1]
    
    
    st_HTML = []
    
    st_cfg = {} #html
    st_cfg_txt = {} #txt
    
    
    for idx, response in enumerate([response1, response2, response3]):
        
        current_week_by_dates = all_dates[idx]
        week_name, days_data = response

        st_cfg[week_name] = {}
        
        
        for day_idx, (dayN, dayC) in enumerate(days_data.items()):
            stringHTML = """"""
            stringTXT = """"""
            # if dayN == "Понедельник":
            #     stringHTML += "<div class='week'>"
            stringTXT += f"\t📆 {dayN} {current_week_by_dates[day_idx] if  day_idx < len(current_week_by_dates) else ""}\n"
            stringHTML += "<div class='day'>"
            stringHTML += f"<br><h2>{dayN}{"<br>"+ current_week_by_dates[day_idx] if  day_idx < len(current_week_by_dates) else ""}</h2>"
            
            if not dayC:
                stringHTML += "<p>Занятий нет</p>"
                stringTXT += "\nЗанятий нет\n"
            else:
                for lesson in dayC:
                    time_code = int(lesson["time_code"])
                    time_range = codes.get(time_code, ("", ""))

                    stringHTML += (
                        f"<h4>{lesson['time']}</h4>"
                        f"<h5>{time_range[0]} - {time_range[1]}</h5>"
                        f"<h6>{lesson['teacher']}</h6>"
                        f"<p>{lesson['subject']} <span>({lesson['room']})</span></p>"
                    )
                    
                    stringTXT += (
                        f"\n{lesson['time']}\n"
                        f"  {time_range[0]} - {time_range[1]}\n"
                        f"  {lesson['teacher']}\n"
                        f"  {lesson['subject']} ({lesson['room']})\n"
                    )
            stringHTML += "</div>"
            stringTXT += "\n"
            
            if day_idx < len(current_week_by_dates):
                date_key = current_week_by_dates[day_idx]
                st_cfg[week_name][date_key] = stringHTML
                

        full_week_html = "<div class='week'>" + "".join(st_cfg[week_name].values()) + "</div>"
        st_HTML.append(full_week_html)

    need_data = """"""
    for week, data in st_cfg.items():
        need_data += "<div class='week'>"
        for key_day, day_data in data.items():
            for d in req_days:
                if d == key_day:
                    need_data += day_data
        need_data += "</div>"
    
    
    style_GREEN = """<style> * { margin: 0; padding: 0; font-family: 'Roboto', serif; } header { background: #85d44f; box-shadow: inset 8px 8px 8px #a3d1ad6c, inset -8px -8px 8px #bae27960; text-align: center; margin: auto; padding: .5rem; position: sticky; transition: background .5s ease, width 1s, transform 1s; backdrop-filter: blur(6px); border-radius: 30px; width: calc(100% - 2rem); top: 0; animation: w 2s forwards; z-index: 1000000; h1 { letter-spacing: 2px; color: #fff; } } body { background: linear-gradient(to bottom, #fdf9f9, #edfff6); } @keyframes w { 100% { transform: translateY(1rem); } } main { display: grid; grid-template-rows: repeat(3, 1fr); } :root { --bg-week: #f4f7f2; --bg-day: #ffffff; --accent-color: #4CAF50; --text-main: #2d3436; --text-secondary: #636e72; --border-color: #e0eadd; --shadow: 0 4px 15px rgba(0, 0, 0, 0.05);} .week { background: var(--bg-week); display: flex; gap: 15px; padding: 2em; margin: 1em auto; border-radius: 20px; overflow-x: auto; } .day { background: var(--bg-day); padding: 1.5em; flex: 1; min-width: 200px; border-radius: 16px; box-shadow: var(--shadow); transition: box-shadow 0.2s ease; border: 1px solid var(--border-color); cursor: pointer; &:hover { box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15); } h2 { color: var(--accent-color); text-align: center; margin: -1.5em -1.5em 1em -1.5em; padding: 1em 0; background: rgba(76, 175, 80, 0.08); border-radius: 16px 16px 0 0; font-size: 1.2em; text-transform: uppercase; letter-spacing: 1.5px; } h4 { color: var(--text-main); font-size: 1rem; margin-top: 1.2em; padding-top: 0.8em; border-top: 1px dashed var(--border-color); display: flex; justify-content: space-between; align-items: center; } h5 { color: var(--text-secondary); font-weight: 400; font-size: 0.9rem; margin-top: 0.3em; line-height: 1.4; } h6 { font-weight: 500; font-size: small; font-style: italic; } } @media (max-width:1440px) { main{ display: flex;} .week{flex-direction:column;} } @media (max-width:1024px) { main {display: grid; grid-template-columns: 1fr !important;} .day {pointer-events:none}} </style>"""    
    script_GREEN = """<script> document.querySelector("header").addEventListener("click", function() {window.scrollTo({top: 0, behavior: "smooth"})}); window.addEventListener("scroll", function(){ if (window.scrollY >= 50) { document.querySelector("header").style.background = "#84d44fa2";document.querySelector("header").style.width = "80%"} else { document.querySelector("header").style.background = "#85d44f"; document.querySelector("header").style.width = "calc(100% - 2rem)" } }); var d=new Date(),n=d.getDay(),m=d.getMonth(),dt=d.getDate(),days={1:"Понедельник",2:"Вторник",3:"Среда",4:"Четверг",5:"Пятница",6:"Суббота",7:"Воскресенье"},months={0:"января",1:"февраля",2:"марта",3:"апреля",4:"мая",5:"июня",6:"июля",7:"августа",8:"сентября",9:"октября",10:"ноября",11:"декабря"},dayCall=days[n]??"Воскресенье",full=dayCall+", "+dt+" "+(months[m]??0);document.getElementById("nowDateFull").innerHTML=full;document.querySelectorAll(".day").forEach(el=>{if(el.innerHTML===dayCall)el.classList.add("now")}); window.addEventListener("DOMContentLoaded", function () {document.querySelectorAll(".week").forEach((w) => {if (w.innerHTML.length == 0) {w.remove()}})})</script>"""
    
    style_BLUE = """<style> * { margin: 0; padding: 0; font-family: 'Roboto', serif; } header { background: #367ed1; box-shadow: inset 8px 8px 12px #14e4ff4d, inset -8px -8px 12px #3717c72d; text-align: center; margin: auto; padding: .5rem; position: sticky; transition: background .5s ease, width 1s, transform 1s; backdrop-filter: blur(6px); border-radius: 30px; width: calc(100% - 2rem); top: 0; animation: w 2s forwards; z-index: 1000000; h1 { letter-spacing: 2px; color: #fff; } } body { background: linear-gradient(to bottom, #f9fcfd, #d4e4dc); } @keyframes w { 100% { transform: translateY(1rem); } } main { display: grid; grid-template-rows: repeat(3, 1fr); } :root { --bg-week: #bed3da; --bg-day: #ffffff; --accent-color: #3f87e6; --text-main: #2d3436; --text-secondary: #527785; --border-color: #b6b1cf; --shadow: 0 4px 15px rgba(0, 0, 0, 0.075); } .week { background: var(--bg-week); display: flex; gap: 15px; padding: 2em; margin: 1em auto; border-radius: 20px; overflow-x: auto; } .day { background: var(--bg-day); padding: 1.5em; flex: 1; min-width: 200px; border-radius: 16px; box-shadow: var(--shadow); transition: box-shadow 0.2s ease; border: 1px solid var(--border-color); cursor: pointer; &:hover { box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15); } h2 { color: var(--accent-color); text-align: center; margin: -1.5em -1.5em 1em -1.5em; padding: 1em 0; background: rgba(76, 111, 175, 0.08); border-radius: 16px 16px 0 0; font-size: 1.2em; text-transform: uppercase; letter-spacing: 1.5px; } h4 { color: var(--text-main); font-size: 1rem; margin-top: 1.2em; padding-top: 0.8em; border-top: 1px dashed var(--border-color); display: flex; justify-content: space-between; align-items: center; } h5 { color: var(--text-secondary); font-weight: 400; font-size: 0.9rem; margin-top: 0.3em; line-height: 1.4; } h6 { font-weight: 500; font-size: small; font-style: italic; } } @media (max-width:1440px) { main { display: flex; } .week { flex-direction: column; } } @media (max-width:1024px) { main { display: grid; grid-template-columns: 1fr 1fr !important; } } @media (max-width:1024px) { main { display: grid; grid-template-columns: 1fr !important; } .day { pointer-events: none } } </style>"""
    script_BLUE = """<script> document.querySelector("header").addEventListener("click", function() {window.scrollTo({top: 0, behavior: "smooth"})});  window.addEventListener("scroll", function () { if (window.scrollY >= 50) { document.querySelector("header").style.background = "#367ed17e"; document.querySelector("header").style.width = "80%" } else { document.querySelector("header").style.background = "#367ed1"; document.querySelector("header").style.width = "calc(100% - 2rem)" } }); var d = new Date(), n = d.getDay(), m = d.getMonth(), dt = d.getDate(), days = { 1: "Понедельник", 2: "Вторник", 3: "Среда", 4: "Четверг", 5: "Пятница", 6: "Суббота", 7: "Воскресенье" }, months = { 0: "января", 1: "февраля", 2: "марта", 3: "апреля", 4: "мая", 5: "июня", 6: "июля", 7: "августа", 8: "сентября", 9: "октября", 10: "ноября", 11: "декабря" }, dayCall = days[n] ?? "Воскресенье", full = dayCall + ", " + dt + " " + (months[m] ?? 0); document.getElementById("nowDateFull").innerHTML = full; document.querySelectorAll(".day").forEach(el => { if (el.innerHTML === dayCall) el.classList.add("now") }); window.addEventListener("DOMContentLoaded", function () { document.querySelectorAll(".week").forEach((w) => { if (w.innerHTML.length == 0) { w.remove() } }) })</script>"""
    
    style_ORANGE = """<style> * { margin: 0; padding: 0; font-family: 'Roboto', serif; } header { background: #ff9800; box-shadow: inset 8px 8px 12px #ffc10766, inset -8px -8px 12px #e651004d; text-align: center; margin: auto; padding: .5rem; position: sticky; transition: background .5s ease, width 1s, transform 1s; backdrop-filter: blur(6px); border-radius: 30px; width: calc(100% - 2rem); top: 0; animation: w 2s forwards; z-index: 1000000; h1 { letter-spacing: 2px; color: #fff; } } body { background: linear-gradient(to bottom, #fffdfa, #ffe0b2); min-height: 100vh; } @keyframes w { 100% { transform: translateY(1rem); } } main { display: grid; grid-template-rows: repeat(3, 1fr); } :root { --bg-week: #ffecb3; --bg-day: #ffffff; --accent-color: #f57c00; --text-main: #4e342e; --text-secondary: #8d6e63; --border-color: #ffe082; --shadow: 0 4px 15px rgba(245, 124, 0, 0.1); } .week { background: var(--bg-week); display: flex; gap: 15px; padding: 2em; margin: 1em auto; border-radius: 20px; overflow-x: auto; } .day { background: var(--bg-day); padding: 1.5em; flex: 1; min-width: 200px; border-radius: 16px; box-shadow: var(--shadow); transition: all 0.3s ease; border: 1px solid var(--border-color); cursor: pointer; &:hover { box-shadow: 0 8px 25px rgba(245, 124, 0, 0.2); transform: translateY(-5px); } h2 { color: var(--accent-color); text-align: center; margin: -1.5em -1.5em 1em -1.5em; padding: 1em 0; background: rgba(255, 152, 0, 0.1); border-radius: 16px 16px 0 0; font-size: 1.2em; text-transform: uppercase; letter-spacing: 1.5px; } h4 { color: var(--text-main); font-size: 1rem; margin-top: 1.2em; padding-top: 0.8em; border-top: 1px dashed var(--border-color); display: flex; justify-content: space-between; align-items: center; } h5 { color: var(--text-secondary); font-weight: 400; font-size: 0.9rem; margin-top: 0.3em; line-height: 1.4; } h6 { font-weight: 500; font-size: small; font-style: italic; } } @media (max-width:1440px) { main { display: flex; } .week { flex-direction: column; } } @media (max-width:1024px) { main { display: grid; grid-template-columns: 1fr 1fr !important; } } @media (max-width:768px) { main { grid-template-columns: 1fr !important; } .day { pointer-events: none } } </style>"""
    script_ORANGE = """<script> document.querySelector("header").addEventListener("click", function() {window.scrollTo({top: 0, behavior: "smooth"})});  window.addEventListener("scroll", function () { const head = document.querySelector("header"); if (window.scrollY >= 50) { head.style.background = "#ff990077"; head.style.width = "85%"; } else { head.style.background = "#ff9800"; head.style.width = "calc(100% - 2rem)"; }}); var d=new Date(),n=d.getDay(),m=d.getMonth(),dt=d.getDate(),days={1:"Понедельник",2:"Вторник",3:"Среда",4:"Четверг",5:"Пятница",6:"Суббота",0:"Воскресенье"},months={0:"января",1:"февраля",2:"марта",3:"апреля",4:"мая",5:"июня",6:"июля",7:"августа",8:"сентября",9:"октября",10:"ноября",11:"декабря"},dayCall=days[n],full=dayCall+", "+dt+" "+months[m];document.getElementById("nowDateFull").innerHTML=full;document.querySelectorAll(".day h2").forEach(el=>{if(el.innerText.trim().toLowerCase()===dayCall.toLowerCase())el.closest(".day").classList.add("now")});window.addEventListener("DOMContentLoaded",function(){document.querySelectorAll(".week").forEach(w=>{if(w.innerText.trim().length==0)w.remove()})});</script>"""
    
    style_DARK = """<style> * { margin: 0; padding: 0; font-family: 'Roboto', sans-serif; } header { background: #1e293b; box-shadow: 0 4px 20px rgba(0,0,0,0.4); text-align: center; margin: auto; padding: .5rem; position: sticky; transition: all .5s ease; backdrop-filter: blur(6px); border-radius: 30px; width: calc(100% - 2rem); top: 0; animation: w 2s forwards; z-index: 1000000; border: 1px solid #334155; h1 { letter-spacing: 3px; color: #38bdf8; text-shadow: 0 0 10px rgba(56, 189, 248, 0.3); } } body { background: #000; color: #f1f5f9; } @keyframes w { 100% { transform: translateY(1rem); } } :root { --bg-week: #1e293b9e; --bg-day: linear-gradient(-45deg, black, transparent); --accent-color: #38bdf8; --text-main: #f1f5f9; --text-secondary: #94a3b8; --border-color: #475569; --shadow: 0 10px 30px rgba(0, 0, 0, 0.5); } .week { background: var(--bg-week); display: flex; gap: 20px; padding: 2em; margin: 1em auto; border-radius: 24px; } .day { background: var(--bg-day); padding: 1.5em; flex: 1; min-width: 220px; border-radius: 18px; box-shadow: var(--shadow); border: 1px solid var(--border-color); &:hover { border-color: var(--accent-color); background: #3d4b5f; } h2 { color: var(--accent-color); text-align: center; margin: -1.5em -1.5em 1em -1.5em; padding: 1em 0; background: rgba(56, 189, 248, 0.1); border-radius: 18px 18px 0 0; font-size: 1.1em; } h4 { color: var(--text-main); border-top: 1px solid var(--border-color); margin-top: 1.2em; padding-top: 0.8em; } h5 { color: var(--text-secondary); }} h6 {font-weight: 200;} .now { border: 2px solid var(--accent-color) !important; box-shadow: 0 0 20px rgba(56, 189, 248, 0.2); } @media (max-width:1440px) { main{ display: flex;} .week{flex-direction:column;} }  @media (max-width:1024px) { main {display: grid; grid-template-columns: 1fr !important;} .day {pointer-events:none}} </style>"""
    script_DARK = """<script> document.querySelector("header").addEventListener("click", function() {window.scrollTo({top: 0, behavior: "smooth"})}); window.addEventListener("scroll", function () { const head = document.querySelector("header"); if (window.scrollY >= 50) { head.style.background = "#1e293bcc"; head.style.width = "70%"; head.style.borderColor = "#38bdf8"; } else { head.style.background = "#1e293b"; head.style.width = "calc(100% - 2rem)"; head.style.borderColor = "#334155"; }});var d=new Date(),n=d.getDay(),m=d.getMonth(),dt=d.getDate(),days={1:"Понедельник",2:"Вторник",3:"Среда",4:"Четверг",5:"Пятница",6:"Суббота",0:"Воскресенье"},months={0:"января",1:"февраля",2:"марта",3:"апреля",4:"мая",5:"июня",6:"июля",7:"августа",8:"сентября",9:"октября",10:"ноября",11:"декабря"},dayCall=days[n],full=dayCall+", "+dt+" "+months[m];document.getElementById("nowDateFull").innerHTML=full;document.querySelectorAll(".day h2").forEach(el=>{if(el.innerText.trim().toLowerCase()===dayCall.toLowerCase())el.closest(".day").classList.add("now")});window.addEventListener("DOMContentLoaded",function(){document.querySelectorAll(".week").forEach(w=>{if(w.innerText.trim().length==0)w.remove()})});</script>"""
    
    style_PURPLE = """<style> * { margin: 0; padding: 0; font-family: 'Roboto', serif; } header { background: #6c5ce7; box-shadow: 0 10px 15px -3px rgba(108, 92, 231, 0.3); text-align: center; margin: auto; padding: .6rem; position: sticky; transition: all .6s cubic-bezier(0.4, 0, 0.2, 1); backdrop-filter: blur(8px); border-radius: 50px; width: calc(100% - 4rem); top: 0; animation: w 2s forwards; z-index: 1000000; h1 { letter-spacing: 4px; color: #fff; font-weight: 300; } } body { background: linear-gradient(135deg, #f3f0ff 0%, #e0d7ff 100%); min-height: 100vh; } :root { --bg-week: rgba(255, 255, 255, 0.5); --bg-day: #ffffff; --accent-color: #6c5ce7; --text-main: #2d3436; --text-secondary: #636e72; --border-color: #dcdde1; --shadow: 0 10px 20px rgba(108, 92, 231, 0.05); } .week { background: var(--bg-week); display: flex; gap: 15px; padding: 2.5em; margin: 1.5em auto; border-radius: 30px; backdrop-filter: blur(10px); border: 1px solid rgba(255,255,255,0.8); } .day { background: var(--bg-day); padding: 1.8em; flex: 1; min-width: 210px; border-radius: 22px; box-shadow: var(--shadow); transition: transform 0.4s ease; &:hover { transform: scale(1.03); box-shadow: 0 15px 30px rgba(108, 92, 231, 0.15); } h2 { color: #fff; background: var(--accent-color); margin: -2em -0 .8em 0; padding: .5em 0; border-radius: 22px 22px 0 0; text-align: center; } h4 { border-top: 2px solid #f3f0ff; padding-top: .5em; }} h5 { font-weight: 400; letter-spacing: 2px } h6 {font-weight: 200;} @keyframes w { 100% { transform: translateY(1.5rem); } } @media (max-width:1440px) { main { display: flex; } .week { flex-direction: column; } } @media (max-width:1024px) { main { display: grid; grid-template-columns: 1fr 1fr !important; } } @media (max-width:768px) { main { grid-template-columns: 1fr !important; } .day { pointer-events: none } }  </style>"""
    script_PURPLE = """<script> document.querySelector("header").addEventListener("click", function() {window.scrollTo({top: 0, behavior: "smooth"})}); window.addEventListener("scroll", function () { const head = document.querySelector("header"); if (window.scrollY >= 50) { head.style.background = "#6c5ce7ee"; head.style.width = "75%"; head.style.borderRadius = "20px"; } else { head.style.background = "#6c5ce7"; head.style.width = "calc(100% - 4rem)"; head.style.borderRadius = "50px"; } }); var d=new Date(),n=d.getDay(),m=d.getMonth(),dt=d.getDate(),days={1:"Понедельник",2:"Вторник",3:"Среда",4:"Четверг",5:"Пятница",6:"Суббота",0:"Воскресенье"},months={0:"января",1:"февраля",2:"марта",3:"апреля",4:"мая",5:"июня",6:"июля",7:"августа",8:"сентября",9:"октября",10:"ноября",11:"декабря"},dayCall=days[n],full=dayCall+", "+dt+" "+months[m];document.getElementById("nowDateFull").innerHTML=full;document.querySelectorAll(".day h2").forEach(el=>{if(el.innerText.trim().toLowerCase()===dayCall.toLowerCase())el.closest(".day").classList.add("now")});window.addEventListener("DOMContentLoaded",function(){document.querySelectorAll(".week").forEach(w=>{if(w.innerText.trim().length==0)w.remove()})});</script>"""
    
    
    if theme == "green":
        style, script = style_GREEN, script_GREEN
    elif theme == "blue":
        style, script = style_BLUE, script_BLUE
    elif theme == "orange":
        style, script = style_ORANGE, script_ORANGE
    elif theme == "dark":
        style, script = style_DARK, script_DARK
    elif theme == "purple":
        style, script = style_PURPLE, script_PURPLE
    
    
    new_HTML = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><title>Расписание</title><link rel="stylesheet" href="https://fonts.googleapis.com/css?family=Roboto">""" + style + """</head><body><header><h1>Расписание</h1><object id="nowDateFull" style="color: #d4ffdf; font-weight: 700; letter-spacing: 1px;"></object></header><main>""" + need_data + "</main></body>" + script +"</html>"

    path = io.BytesIO()
    path.write(new_HTML.encode("utf-8"))
    path.seek(0)
    await callback.answer("Файл готов. Отправка...")
    await callback.message.answer_document(
        document=types.BufferedInputFile(
                path.getvalue(),
                filename=f"schedule_{group_name}.html"
            ),
            caption=f"Расписание для {group_name}\n\nСтиль — <b>{theme}</b>",
            parse_mode="HTML"
        )
    await callback.answer("Скачано!")
    done = True
    
    if done:
        if state:
            await state.clear()
